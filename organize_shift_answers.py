import openpyxl
import pprint

#######　←change these places according to the file

path = "answer_example.xlsx"                             #######


startDayColumn = 7  #begin with 1                                        #######
endDayColumn = 13                                                        #######

wb = openpyxl.load_workbook(path)
sheet = wb['フォームの回答 1']
labelList = []

rowIndex=1 #Index of row
columnIndex=1 #index of column
#sheet.cell's columns and rows begins with number 1

numAllRows = sheet.max_row #number of all row
numAllColumns = sheet.max_column #number of all column


# answers of members
tmp = sheet.iter_rows(min_row=2, max_row=numAllRows, min_col=1, max_col=numAllColumns)
answerList = [[cell.value for cell in row] for row in tmp]
#pprint.pprint(answerList)

# lists of spreadsheet column-label
for columnIndex in range(1,numAllColumns):
    labelList.append((sheet.cell(row = 1, column = columnIndex).value))
#pprint.pprint(labelList)   #for check


# declare Lists and Dictionary for available member at each satellite pass
PassLists = ['1', '2', '3', '4', '5', '6', '7', '8'] ######
shiftDicts = {}
keysLists = []

i = 0
for rowIndex in range(startDayColumn - 1, endDayColumn):
    keysLists.append(labelList[rowIndex])
    i = i + 1


for keysList in keysLists:
    shiftDicts[keysList] = {}


for dayIndex in range (startDayColumn-1,endDayColumn):              
    #pprint.pprint(answerList[rowIndex][dayIndex])
    for PassListsIndex in range(len(PassLists)):
        shiftDicts[labelList[dayIndex]][PassLists[PassListsIndex]] = []
        for rowIndex in range(0,numAllRows-1):
            if PassLists[PassListsIndex] in str(answerList[rowIndex][dayIndex]):
                shiftDicts[labelList[dayIndex]][PassLists[PassListsIndex]].append\
                    (str(answerList[rowIndex][2]))


pprint.pprint(shiftDicts)


sheet_new = wb.create_sheet('shift_sheet')                                 #######

for dayIndex in range (startDayColumn-1,endDayColumn):
    sheet_new.cell(row=1, column = dayIndex - startDayColumn + 2, \
            value = keysLists[dayIndex - startDayColumn + 1])
    j=1
    for PassListsIndex in range(len(PassLists)):
        sheet_new.cell(row=PassListsIndex + 1 + j, \
            column = dayIndex - startDayColumn + 2, \
                value = PassLists[PassListsIndex])    # write hour-label

        for i in range(len(shiftDicts[labelList[dayIndex]][PassLists[PassListsIndex]])):
            sheet_new.cell(row=PassListsIndex + 1+i+j+1, column = dayIndex - startDayColumn + 2, \
                value = shiftDicts[labelList[dayIndex]][PassLists[PassListsIndex]][i])
        j=j+numAllRows+1  # +1 for the space of hour-label


target_row = 1
delete_boolean = True #if Ture, delete　
while target_row < sheet_new.max_row:
    delete_boolean = True
    
    for col in range(1,len(keysLists)+1):
        
        if not sheet_new.cell(row = target_row, column = col).value is None:
            target_row = target_row + 1
            delete_boolean = False
            break
    
    if delete_boolean == True:
        sheet_new.delete_rows(target_row)
        print("delete")
    




path = path[0:-9] + '_arranged' + '.xlsx'

wb.save(path)                                 #######

